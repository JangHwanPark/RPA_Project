from openpyxl import load_workbook
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)


#속성 = data_only
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가져옴
for row in ws.values:
    for cell in row:
        print(cell)

# 데이터를 가져올때 수식을 계산하지는 않음
# evaluate 되지 않은 상태의 데이터는 None으로 표시

# 해당 xlsx을 열어서 저장후 실행시 계산된(수식)값이 노출된다.