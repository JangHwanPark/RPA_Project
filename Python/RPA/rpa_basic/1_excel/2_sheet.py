from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 Sheet 디폴트 네임으로 생성
ws.title = "MySheet" # Sheet이름 변경
ws.sheet_properties.tabColor = "ff66ff" #RGB형태로 값을 넣어주면 탭 색상 변경

ws1 = wb.create_sheet("YourSheet") #주어진 흐름으로 시트 생성
ws2 = wb.create_sheet("NewSheet", 2) #2번째 index에 sheet생성

new_ws = wb["NewSheet"] # dict형태로 Sheet에 접근
print(wb.sheetnames) # 모든 시트 이름 확인

# Sheet복사
new_ws["A1"] = "Test" #데이터 입력
target = wb.copy_worksheet(new_ws)
target.title = "Copid Sheet"

wb.save("sample.xlsx")