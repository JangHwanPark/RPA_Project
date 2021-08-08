# 성적표
from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 입력
ws.append(["번호", "영어", "수학"]) # A, B, C
for i in range(1, 11): # 10개 데이터 입력
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 영어 column만 가져오기
print(col_B)
for cell in col_B:
    print(cell.value)

col_range = ws["B:C"] # B ~ C 까지의 column값을 가져옴
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1] # 1번째 row만 가지고 오기
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] # 1번째 줄인 title을 제외하고 2번째 줄에서 6번째 줄에서 가져오기(6포함)
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

from openpyxl.utils.cell import coordinate_from_string, coordinate_to_tuple

row_range_n2 = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range_n2:
    for cell in rows:
        #print(cell.value, end=" ")
        #print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate) #튜플 형태로 끊어줌
        #print(xy, end=" ")
        print(xy[0], end="")
        print(xy[1], end=" ")
    print()

# 전체 rows
#print(tuple(ws.rows)) #한 줄
for row in tuple(ws.rows):
    print(row[1].value) # row값중 첫번째 index cell에 접근하여 출력한다(value)

# 전체 columns
#print(tuple(ws.columns)) #한 열
for column in tuple(ws.columns):
    print(column[0].value) # row랑 동일

for row in ws.iter_rows(): # 전체 row에 대해 반복하면서 출력
    print(row[2].value)

for column in ws.iter_cols(): # 전체 column에 대해 반복하면서 출력
    print(column[0].value)

print("\n--- 구분선 ---\n")

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    #print(row[0].value, row[1].value) # 수학, 영어
    print(row)

for col in ws.itercols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)
# min/max_row값은 지정하지않으면 최대/최솟값을 출력함

wb.save("sample.xlsx")