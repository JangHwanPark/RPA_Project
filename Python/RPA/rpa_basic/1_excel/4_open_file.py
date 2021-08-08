from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook("sample.xlsx") # sample.xlsx파일에서 wb를 불러옴
ws = wb.active # 현재 워크북의 활성화된 시트

# cell 데이터 불러오기 ( row / column을 알때 )
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ")# 1 2 3 4 ...자동 줄바꿈
    print()

# cell 갯수를 모를때
for x in range(1, ws.max_row + 1): # 최대 row수
    for y in range(1, ws.max_column + 1): # 최대 column 수
        print(ws.cell(row=x, column=y).value, end=" ")# 1 2 3 4 ...자동 줄바꿈
    print()