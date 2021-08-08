from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1셀에 1이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1셀의 정보를 출력
print(ws["A1"].value) # A1셀의 값을 출력
print(ws["A10"].value) # 값이 없을때는 none을 출력

#row = 1, 2, 3...
#column = A(1), B(2), C(3), ...
print(ws.cell(row=1, column=1).value) # ws.["A1"].value ==
print(ws.cell(row=1, column=2).value) # ws.["B1"].value ==
# row와 column의 입력 순서는 상관없다.

C = ws.cell(row=1, column=3, value=10) # ws.["C1"].value= 10
print(C.value)

from random import*
index = 1
# 반복문을 이용해서 랜덤 숫자 채우기
for x in range(1, 11): # 10개 row값
    for y in range(1, 11): # 10개 column값
        #ws.cell(row=x, column=y, value=randint(0,100)) # 0~100사이의 숫자 random값
        ws.cell(row = x, column = y, value=index)
        index += 1
wb.save("sample.xlsx")